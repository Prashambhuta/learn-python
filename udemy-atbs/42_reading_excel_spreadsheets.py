#! usr/bin/env/python3
"""
Reading data from excel spreadsheets.
Openoffice & Libreoffice will also work but not .ods format.
"""

"""
Let's use the /vocab/my-word.ods for learning.
"""

# first install 3rd party module openpyxl (open python excel)
# `pip3 install openpyxl`

# import modules
import openpyxl
import os

# get the workbook directory
os.chdir('/home/prasham/vocab')

# create a workbook object by loading the workbook
workbook = openpyxl.open('my-word.xlsx')

""" # IMP!
to load .ods file use module `pyexcel-ods`
"""

# generate a sheet object by using `get_sheet_by_name`
sheet = workbook.get_sheet_by_name('Sheet1')

# get cells and the value of cell
print(sheet['A1'].value)    # returns the value of A1 cell

# get cell id using method(.cell)
print(sheet.cell(1, 1))     # returns cell ID
print(sheet.cell(row=1, column=1).value)   # returns the value of the cell id

# get range of values
# eg. get top 10 rows and columns of worksheet
for i in range(1, 10):
    print(i, sheet.cell(i, 1).value,  "\t", sheet.cell(i, 2).value)






